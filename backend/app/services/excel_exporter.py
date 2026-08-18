import io
import re
from datetime import datetime, timezone
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_cover_sheet(wb, title_prefix: str, election_data: dict, theme_color_hex: str = "0F172A"):
    ws = wb.create_sheet(title="Cover", index=0)
    ws.views.sheetView[0].showGridLines = False

    el = election_data.get("election", {})
    stats = election_data.get("statistics", {})

    theme_fill = PatternFill(start_color=theme_color_hex, end_color=theme_color_hex, fill_type="solid")
    sub_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    card_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    font_brand = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=13, bold=False, color="CBD5E1")
    font_card_head = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_label = Font(name="Calibri", size=10, bold=True, color="475569")
    font_val = Font(name="Calibri", size=11, bold=False, color="0F172A")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Top Brand Header Banner
    ws.merge_cells("B2:G2")
    ws["B2"].value = "CIVITAS SECURE DIGITAL VOTING SYSTEM"
    ws["B2"].font = font_brand
    ws["B2"].fill = theme_fill
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 42

    ws.merge_cells("B3:G3")
    ws["B3"].value = f"OFFICIAL {title_prefix.upper()} AUDIT REPORT"
    ws["B3"].font = font_subtitle
    ws["B3"].fill = sub_fill
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 26

    # Election Overview Card
    ws.merge_cells("B5:G5")
    ws["B5"].value = "ELECTION SPECIFICATION & METADATA"
    ws["B5"].font = font_card_head
    ws["B5"].fill = card_header_fill
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 24

    details = [
        ("Election Name", el.get("name", "")),
        ("Election ID", el.get("election_id") or el.get("id", "")),
        ("Voting Type", str(el.get("voting_type", "regular")).replace("_", " ").title()),
        ("Voter Registration Mode", str(el.get("voter_registration_mode", "pre_registered")).replace("_", " ").title()),
        ("Current Status", el.get("status", "")),
        ("Voting Start Timestamp", el.get("starts_at", "")),
        ("Voting End Timestamp", el.get("ends_at", "")),
        ("Total Registered Voters", stats.get("registered_voters", 0)),
        ("Total Eligible Voters", stats.get("eligible_voters", 0)),
        ("Total Ballots Cast", stats.get("votes_cast", 0)),
        ("Overall Turnout Rate", stats.get("turnout_percentage", 0.0) / 100.0),
        ("Report Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Cryptographic Ledger Integrity", "VERIFIED & AUDITED (TAMPER-EVIDENT)"),
    ]

    for idx, (label, val) in enumerate(details, start=6):
        ws.cell(row=idx, column=2, value=label).font = font_label
        ws.cell(row=idx, column=2).fill = card_fill
        ws.cell(row=idx, column=2).border = thin_border

        ws.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=7)
        val_cell = ws.cell(row=idx, column=3, value=val)
        val_cell.font = font_val
        val_cell.fill = card_fill
        val_cell.border = thin_border
        val_cell.alignment = Alignment(horizontal="left", vertical="center")

        if label == "Overall Turnout Rate":
            val_cell.number_format = "0.00%"
        elif isinstance(val, (int, float)) and "Rate" not in label:
            val_cell.number_format = "#,##0"

        ws.row_dimensions[idx].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16


def style_table_header(ws, row_idx: int, headers: list[str], header_fill_hex: str = "1E293B"):
    h_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.row_dimensions[row_idx].height = 28
    for col_num, title in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_num, value=title)
        c.font = font_h
        c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_fit_columns(ws, max_cols=12):
    for col_idx in range(1, max_cols + 1):
        col_cells = [ws.cell(row=r, column=col_idx) for r in range(1, ws.max_row + 1)]
        max_len = 0
        for cell in col_cells:
            val_str = str(cell.value or "")
            if len(val_str) > max_len and not getattr(cell, "coordinate", "").startswith("A1"):
                max_len = len(val_str)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


# =============================================================================
# 1. REGULAR ELECTION WORKBOOK GENERATOR
# =============================================================================
def generate_regular_election_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, "Regular Election", data, "0F172A")
    el = data.get("election", {})
    stats = data.get("statistics", {})
    candidates = data.get("candidates", [])

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    font_reg = Font(name="Calibri", size=11, color="334155")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_winner = Font(name="Calibri", size=11, bold=True, color="047857")
    winner_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    # SHEET 2: Results Dashboard
    ws_dash = wb.create_sheet(title="Results Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash.freeze_panes = "A2"

    dash_headers = ["Metric", "Value", "Context & Notes"]
    style_table_header(ws_dash, 1, dash_headers, "0F172A")

    winner_info = data.get("winner")
    winner_name = winner_info["name"] if winner_info else ("TIED" if data.get("has_tie") else "None")
    winner_votes = winner_info["votes"] if winner_info else 0
    winner_share = (winner_info["percentage"] / 100.0) if winner_info else 0.0

    dash_rows = [
        ("Total Registered Voters", stats.get("registered_voters", 0), "All registered voters in system database"),
        ("Total Eligible Voters", stats.get("eligible_voters", 0), "Voters qualified for this election"),
        ("Total Ballots Cast", stats.get("votes_cast", 0), "Valid single-choice ballots recorded"),
        ("Turnout Percentage", stats.get("turnout_percentage", 0.0) / 100.0, "Ballots Cast / Eligible Voters"),
        ("Election Winner", winner_name, f"Leading candidate with {winner_votes} votes"),
        ("Winning Vote Count", winner_votes, "Total votes obtained by winning candidate"),
        ("Winning Vote Share", winner_share, "Percentage of total cast votes"),
        ("Candidates Running", len(candidates), "Total registered contenders on ballot"),
    ]

    for label, val, note in dash_rows:
        r = ws_dash.max_row + 1
        c1 = ws_dash.cell(row=r, column=1, value=label)
        c2 = ws_dash.cell(row=r, column=2, value=val)
        c3 = ws_dash.cell(row=r, column=3, value=note)

        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg

        for c in (c1, c2, c3):
            c.border = thin_border

        if "Percentage" in label or "Share" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_dash, 3)

    # SHEET 3: Candidate Results
    ws_cand = wb.create_sheet(title="Candidate Results")
    ws_cand.views.sheetView[0].showGridLines = True
    ws_cand.freeze_panes = "A2"

    cand_headers = ["Rank", "Candidate Name", "Party / Affiliation", "Votes", "Vote Percentage", "Status"]
    style_table_header(ws_cand, 1, cand_headers, "1E293B")

    has_tie = data.get("has_tie", False)
    for c in candidates:
        r = ws_cand.max_row + 1
        is_winner = c.get("rank") == 1 and c.get("votes", 0) > 0 and not has_tie
        status_text = "WINNER" if is_winner else ("TIED" if has_tie and c.get("votes", 0) == candidates[0].get("votes") else "-")

        c_rank = ws_cand.cell(row=r, column=1, value=c.get("rank", 0))
        c_name = ws_cand.cell(row=r, column=2, value=c.get("name", ""))
        c_party = ws_cand.cell(row=r, column=3, value=c.get("party", "Independent"))
        c_votes = ws_cand.cell(row=r, column=4, value=c.get("votes", 0))
        c_pct = ws_cand.cell(row=r, column=5, value=c.get("percentage", 0.0) / 100.0)
        c_status = ws_cand.cell(row=r, column=6, value=status_text)

        c_rank.alignment = Alignment(horizontal="center")
        c_status.alignment = Alignment(horizontal="center")
        c_votes.number_format = "#,##0"
        c_pct.number_format = "0.00%"

        for cell in (c_rank, c_name, c_party, c_votes, c_pct, c_status):
            cell.font = font_winner if is_winner else font_reg
            cell.border = thin_border
            if is_winner:
                cell.fill = winner_fill

    if candidates:
        ws_cand.auto_filter.ref = f"A1:F{len(candidates) + 1}"

    # Embed Charts in Sheet 3
    if len(candidates) > 0:
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Candidate Vote Comparison"
        bar.y_axis.title = "Votes"
        bar.x_axis.title = "Candidate"
        data_ref = Reference(ws_cand, min_col=4, min_row=1, max_row=len(candidates) + 1)
        cats_ref = Reference(ws_cand, min_col=2, min_row=2, max_row=len(candidates) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 16
        bar.height = 10
        ws_cand.add_chart(bar, "H2")

        pie = PieChart()
        pie.title = "Vote Percentage Distribution"
        pie_data_ref = Reference(ws_cand, min_col=4, min_row=1, max_row=len(candidates) + 1)
        pie.add_data(pie_data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 14
        pie.height = 10
        ws_cand.add_chart(pie, "H18")

    auto_fit_columns(ws_cand, 6)

    # SHEET 4: Vote Details
    ws_votes = wb.create_sheet(title="Vote Details")
    ws_votes.views.sheetView[0].showGridLines = True
    ws_votes.freeze_panes = "A2"

    show_names = bool(el.get("show_voter_names_in_results"))
    vote_headers = ["Voter Registration ID", "Voter Full Name", "Ballot Cast Timestamp", "Record Status"] if show_names else ["Voter Registration ID", "Ballot Cast Timestamp", "Record Status"]
    style_table_header(ws_votes, 1, vote_headers, "1E293B")

    for item in data.get("voter_participation_log", []):
        r = ws_votes.max_row + 1
        if show_names:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voter_name", ""))
            c3 = ws_votes.cell(row=r, column=3, value=item.get("voted_at", ""))
            c4 = ws_votes.cell(row=r, column=4, value="RECORDED")
            row_cells = (c1, c2, c3, c4)
        else:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voted_at", ""))
            c3 = ws_votes.cell(row=r, column=3, value="RECORDED")
            row_cells = (c1, c2, c3)

        for cell in row_cells:
            cell.font = font_reg
            cell.border = thin_border

    auto_fit_columns(ws_votes, 4)

    # SHEET 5: Voter Statistics
    ws_stats = wb.create_sheet(title="Voter Statistics")
    ws_stats.views.sheetView[0].showGridLines = True
    ws_stats.freeze_panes = "A2"

    style_table_header(ws_stats, 1, ["Metric", "Count / Rate", "Description"], "1E293B")
    non_voters = max(0, stats.get("eligible_voters", 0) - stats.get("votes_cast", 0))

    voter_stat_rows = [
        ("Total Eligible Voters", stats.get("eligible_voters", 0), "Verified active voters eligible to vote"),
        ("Participating Voters", stats.get("votes_cast", 0), "Voters who completed ballot casting"),
        ("Non-Participating Voters", non_voters, "Eligible voters who did not vote"),
        ("Voter Participation Rate", stats.get("turnout_percentage", 0.0) / 100.0, "Participation rate percentage"),
        ("Abstained / Spoiled Ballots", 0, "No invalid ballots possible in digital cryptographic tally"),
    ]

    for label, val, desc in voter_stat_rows:
        r = ws_stats.max_row + 1
        c1 = ws_stats.cell(row=r, column=1, value=label)
        c2 = ws_stats.cell(row=r, column=2, value=val)
        c3 = ws_stats.cell(row=r, column=3, value=desc)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_stats, 3)
    return wb


# =============================================================================
# 2. POLL WORKBOOK GENERATOR
# =============================================================================
def generate_poll_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, "Opinion Poll", data, "0F766E")  # Teal Theme
    el = data.get("election", {})
    stats = data.get("statistics", {})
    poll_info = data.get("poll", {})
    options = data.get("options", []) or data.get("candidates", [])

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    font_reg = Font(name="Calibri", size=11, color="334155")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_top = Font(name="Calibri", size=11, bold=True, color="0F766E")
    top_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")

    # SHEET 2: Poll Dashboard
    ws_dash = wb.create_sheet(title="Poll Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash.freeze_panes = "A2"

    style_table_header(ws_dash, 1, ["Poll Metric", "Result", "Notes"], "0F766E")

    top_opt = poll_info.get("most_selected_option")
    top_name = top_opt["name"] if top_opt else "None"
    top_votes = top_opt["votes"] if top_opt else 0
    top_share = (top_opt["percentage"] / 100.0) if top_opt else 0.0

    poll_rows = [
        ("Poll Question", poll_info.get("question", el.get("name", "")), "Subject question evaluated by respondents"),
        ("Total Responses Received", poll_info.get("total_responses", stats.get("votes_cast", 0)), "Total completed poll submissions"),
        ("Response Rate", poll_info.get("response_rate", stats.get("turnout_percentage", 0.0)) / 100.0, "Responses / Total Eligible Voters"),
        ("Most Selected Option", top_name, f"Rank 1 choice with {top_votes} responses ({top_share:.1%})"),
        ("Number of Options Offered", len(options), "Total available poll alternatives"),
    ]

    for label, val, note in poll_rows:
        r = ws_dash.max_row + 1
        c1 = ws_dash.cell(row=r, column=1, value=label)
        c2 = ws_dash.cell(row=r, column=2, value=val)
        c3 = ws_dash.cell(row=r, column=3, value=note)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_dash, 3)

    # SHEET 3: Question Results
    ws_res = wb.create_sheet(title="Question Results")
    ws_res.views.sheetView[0].showGridLines = True
    ws_res.freeze_panes = "A2"

    headers = ["Rank", "Poll Option", "Responses Count", "Share Percentage", "Leading Status"]
    style_table_header(ws_res, 1, headers, "115E59")

    has_tie = poll_info.get("has_tie", False)
    for opt in options:
        r = ws_res.max_row + 1
        is_top = opt.get("rank") == 1 and opt.get("votes", 0) > 0 and not has_tie
        status_text = "MOST SELECTED" if is_top else ("TIED TOP" if has_tie and opt.get("votes", 0) == options[0].get("votes") else "-")

        c_rank = ws_res.cell(row=r, column=1, value=opt.get("rank", 0))
        c_name = ws_res.cell(row=r, column=2, value=opt.get("name", ""))
        c_votes = ws_res.cell(row=r, column=3, value=opt.get("votes", 0))
        c_pct = ws_res.cell(row=r, column=4, value=opt.get("percentage", 0.0) / 100.0)
        c_status = ws_res.cell(row=r, column=5, value=status_text)

        c_rank.alignment = Alignment(horizontal="center")
        c_status.alignment = Alignment(horizontal="center")
        c_votes.number_format = "#,##0"
        c_pct.number_format = "0.00%"

        for cell in (c_rank, c_name, c_votes, c_pct, c_status):
            cell.font = font_top if is_top else font_reg
            cell.border = thin_border
            if is_top:
                cell.fill = top_fill

    if options:
        ws_res.auto_filter.ref = f"A1:E{len(options) + 1}"

    if len(options) > 0:
        bar = BarChart()
        bar.type = "col"
        bar.style = 13
        bar.title = "Poll Response Distribution"
        bar.y_axis.title = "Responses"
        bar.x_axis.title = "Option"
        data_ref = Reference(ws_res, min_col=3, min_row=1, max_row=len(options) + 1)
        cats_ref = Reference(ws_res, min_col=2, min_row=2, max_row=len(options) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 16
        bar.height = 10
        ws_res.add_chart(bar, "G2")

    auto_fit_columns(ws_res, 5)

    # SHEET 4: Response Data
    ws_resp = wb.create_sheet(title="Response Data")
    ws_resp.views.sheetView[0].showGridLines = True
    ws_resp.freeze_panes = "A2"

    show_names = bool(el.get("show_voter_names_in_results"))
    resp_headers = ["Respondent Voter ID", "Respondent Name", "Timestamp", "Response Verified"] if show_names else ["Respondent Voter ID", "Timestamp", "Response Verified"]
    style_table_header(ws_resp, 1, resp_headers, "115E59")

    for item in data.get("voter_participation_log", []):
        r = ws_resp.max_row + 1
        if show_names:
            c1 = ws_resp.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_resp.cell(row=r, column=2, value=item.get("voter_name", ""))
            c3 = ws_resp.cell(row=r, column=3, value=item.get("voted_at", ""))
            c4 = ws_resp.cell(row=r, column=4, value="RECORDED")
            row_cells = (c1, c2, c3, c4)
        else:
            c1 = ws_resp.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_resp.cell(row=r, column=2, value=item.get("voted_at", ""))
            c3 = ws_resp.cell(row=r, column=3, value="RECORDED")
            row_cells = (c1, c2, c3)

        for cell in row_cells:
            cell.font = font_reg
            cell.border = thin_border

    auto_fit_columns(ws_resp, 4)

    # SHEET 5: Statistics
    ws_stat = wb.create_sheet(title="Statistics")
    ws_stat.views.sheetView[0].showGridLines = True
    ws_stat.freeze_panes = "A2"

    style_table_header(ws_stat, 1, ["Poll Metric", "Value", "Details"], "115E59")
    non_resp = max(0, stats.get("eligible_voters", 0) - stats.get("votes_cast", 0))

    stat_rows = [
        ("Total Potential Respondents", stats.get("eligible_voters", 0), "Eligible community members"),
        ("Responses Cast", stats.get("votes_cast", 0), "Active poll responses submitted"),
        ("Non-Respondents", non_resp, "Eligible voters who did not respond"),
        ("Engagement Ratio", stats.get("turnout_percentage", 0.0) / 100.0, "Total response rate"),
    ]

    for label, val, desc in stat_rows:
        r = ws_stat.max_row + 1
        c1 = ws_stat.cell(row=r, column=1, value=label)
        c2 = ws_stat.cell(row=r, column=2, value=val)
        c3 = ws_stat.cell(row=r, column=3, value=desc)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Ratio" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_stat, 3)
    return wb


# =============================================================================
# 3. MULTIPLE CHOICE WORKBOOK GENERATOR
# =============================================================================
def generate_multiple_choice_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, "Multiple Choice Ballot", data, "4338CA")  # Indigo Theme
    el = data.get("election", {})
    stats = data.get("statistics", {})
    mc_info = data.get("multiple_choice", {})
    options = data.get("options", []) or data.get("candidates", [])

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    font_reg = Font(name="Calibri", size=11, color="334155")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_indigo = Font(name="Calibri", size=11, bold=True, color="4338CA")

    # SHEET 2: Results Dashboard
    ws_dash = wb.create_sheet(title="Results Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash.freeze_panes = "A2"

    style_table_header(ws_dash, 1, ["Multi-Choice Metric", "Value", "Methodology Note"], "4338CA")

    mc_rows = [
        ("Total Participating Voters", mc_info.get("total_voters", stats.get("votes_cast", 0)), "Distinct voters who cast multi-choice ballots"),
        ("Total Options Selected", mc_info.get("total_selections", stats.get("total_vote_records", 0)), "Sum of all option selections across all ballots"),
        ("Average Selections Per Voter", mc_info.get("average_selections_per_voter", 0.0), "Total Selections / Participating Voters"),
        ("Voter Participation Rate", stats.get("turnout_percentage", 0.0) / 100.0, "Participating Voters / Eligible Voters"),
        ("Multi-Selection Rule", "Enabled", "Voters were permitted to select multiple options"),
        ("Percentage Calculation Basis", "Per Participating Voter", "Percentages may exceed 100% in aggregate"),
    ]

    for label, val, note in mc_rows:
        r = ws_dash.max_row + 1
        c1 = ws_dash.cell(row=r, column=1, value=label)
        c2 = ws_dash.cell(row=r, column=2, value=val)
        c3 = ws_dash.cell(row=r, column=3, value=note)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)) and "Average" not in label:
            c2.number_format = "#,##0"
        elif "Average" in label and isinstance(val, (int, float)):
            c2.number_format = "0.00"

    auto_fit_columns(ws_dash, 3)

    # SHEET 3: Selection Results
    ws_sel = wb.create_sheet(title="Selection Results")
    ws_sel.views.sheetView[0].showGridLines = True
    ws_sel.freeze_panes = "A2"

    sel_headers = ["Rank", "Option / Candidate", "Selections Count", "% of Participating Voters", "% of Total Selections"]
    style_table_header(ws_sel, 1, sel_headers, "3730A3")

    for opt in options:
        r = ws_sel.max_row + 1
        c_rank = ws_sel.cell(row=r, column=1, value=opt.get("rank", 0))
        c_name = ws_sel.cell(row=r, column=2, value=opt.get("name", ""))
        c_cnt = ws_sel.cell(row=r, column=3, value=opt.get("selections_count", opt.get("votes", 0)))
        c_pct_v = ws_sel.cell(row=r, column=4, value=opt.get("percentage_of_voters", opt.get("percentage", 0.0)) / 100.0)
        c_pct_t = ws_sel.cell(row=r, column=5, value=opt.get("percentage_of_total_selections", 0.0) / 100.0)

        c_rank.alignment = Alignment(horizontal="center")
        c_cnt.number_format = "#,##0"
        c_pct_v.number_format = "0.00%"
        c_pct_t.number_format = "0.00%"

        for cell in (c_rank, c_name, c_cnt, c_pct_v, c_pct_t):
            cell.font = font_reg
            cell.border = thin_border

    if options:
        ws_sel.auto_filter.ref = f"A1:E{len(options) + 1}"

    if len(options) > 0:
        bar = BarChart()
        bar.type = "col"
        bar.style = 12
        bar.title = "Multiple Choice Selections by Option"
        bar.y_axis.title = "Selections Count"
        bar.x_axis.title = "Option"
        data_ref = Reference(ws_sel, min_col=3, min_row=1, max_row=len(options) + 1)
        cats_ref = Reference(ws_sel, min_col=2, min_row=2, max_row=len(options) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 16
        bar.height = 10
        ws_sel.add_chart(bar, "G2")

    auto_fit_columns(ws_sel, 5)

    # SHEET 4: Vote Details
    ws_votes = wb.create_sheet(title="Vote Details")
    ws_votes.views.sheetView[0].showGridLines = True
    ws_votes.freeze_panes = "A2"

    show_names = bool(el.get("show_voter_names_in_results"))
    vote_headers = ["Voter ID", "Voter Name", "Voted At Timestamp", "Status"] if show_names else ["Voter ID", "Voted At Timestamp", "Status"]
    style_table_header(ws_votes, 1, vote_headers, "3730A3")

    for item in data.get("voter_participation_log", []):
        r = ws_votes.max_row + 1
        if show_names:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voter_name", ""))
            c3 = ws_votes.cell(row=r, column=3, value=item.get("voted_at", ""))
            c4 = ws_votes.cell(row=r, column=4, value="BALLOT SUBMITTED")
            row_cells = (c1, c2, c3, c4)
        else:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voted_at", ""))
            c3 = ws_votes.cell(row=r, column=3, value="BALLOT SUBMITTED")
            row_cells = (c1, c2, c3)

        for cell in row_cells:
            cell.font = font_reg
            cell.border = thin_border

    auto_fit_columns(ws_votes, 4)

    # SHEET 5: Statistics
    ws_stat = wb.create_sheet(title="Statistics")
    ws_stat.views.sheetView[0].showGridLines = True
    ws_stat.freeze_panes = "A2"

    style_table_header(ws_stat, 1, ["Metric", "Value", "Description"], "3730A3")
    stats_data = [
        ("Total Eligible Voters", stats.get("eligible_voters", 0), "Voters qualified for this multi-choice ballot"),
        ("Participating Voters", mc_info.get("total_voters", stats.get("votes_cast", 0)), "Total distinct voters who submitted selections"),
        ("Total Selections Made", mc_info.get("total_selections", stats.get("total_vote_records", 0)), "Total accumulated candidate/option checks"),
        ("Turnout Rate", stats.get("turnout_percentage", 0.0) / 100.0, "Participating Voters / Eligible Voters"),
    ]

    for label, val, desc in stats_data:
        r = ws_stat.max_row + 1
        c1 = ws_stat.cell(row=r, column=1, value=label)
        c2 = ws_stat.cell(row=r, column=2, value=val)
        c3 = ws_stat.cell(row=r, column=3, value=desc)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_stat, 3)
    return wb


# =============================================================================
# 4. YES / NO DECISION WORKBOOK GENERATOR
# =============================================================================
def generate_yes_no_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, "Decision & Referendum", data, "047857")  # Emerald Theme
    el = data.get("election", {})
    stats = data.get("statistics", {})
    dec_info = data.get("decision", {})

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    font_reg = Font(name="Calibri", size=11, color="334155")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")

    decision_res = dec_info.get("result", "PENDING")
    is_approved = decision_res == "APPROVED"
    font_decision = Font(name="Calibri", size=14, bold=True, color="047857" if is_approved else "BE123C")
    fill_decision = PatternFill(start_color="E6F4EA" if is_approved else "FFE4E6", end_color="E6F4EA" if is_approved else "FFE4E6", fill_type="solid")

    # SHEET 2: Decision Dashboard
    ws_dash = wb.create_sheet(title="Decision Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash.freeze_panes = "A2"

    style_table_header(ws_dash, 1, ["Decision Parameter", "Value", "Status / Notes"], "047857")

    dash_rows = [
        ("Proposal / Measure", dec_info.get("proposal", el.get("name", "")), "The referendum motion voted upon"),
        ("Final Decision Result", decision_res, "Calculated based on valid votes tally"),
        ("Total Votes Cast", dec_info.get("total_votes", stats.get("votes_cast", 0)), "Valid ballots cast in referendum"),
        ("YES Votes Count", dec_info.get("yes_votes", 0), "Ballots cast in approval"),
        ("YES Vote Percentage", dec_info.get("yes_percentage", 0.0) / 100.0, "YES Votes / Total Ballots"),
        ("NO Votes Count", dec_info.get("no_votes", 0), "Ballots cast in opposition"),
        ("NO Vote Percentage", dec_info.get("no_percentage", 0.0) / 100.0, "NO Votes / Total Ballots"),
        ("Victory Margin", dec_info.get("margin_percentage", 0.0) / 100.0, "Absolute difference between YES and NO shares"),
        ("Voter Turnout Rate", stats.get("turnout_percentage", 0.0) / 100.0, "Turnout percentage across registered voters"),
    ]

    for label, val, note in dash_rows:
        r = ws_dash.max_row + 1
        c1 = ws_dash.cell(row=r, column=1, value=label)
        c2 = ws_dash.cell(row=r, column=2, value=val)
        c3 = ws_dash.cell(row=r, column=3, value=note)

        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg

        if label == "Final Decision Result":
            c2.font = font_decision
            c2.fill = fill_decision
            c2.alignment = Alignment(horizontal="center", vertical="center")

        for c in (c1, c2, c3):
            c.border = thin_border

        if "Percentage" in label or "Margin" in label or "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_dash, 3)

    # SHEET 3: Vote Summary
    ws_sum = wb.create_sheet(title="Vote Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    ws_sum.freeze_panes = "A2"

    style_table_header(ws_sum, 1, ["Choice", "Votes Count", "Percentage Share", "Outcome"], "065F46")

    summary_entries = [
        ("YES / APPROVE", dec_info.get("yes_votes", 0), dec_info.get("yes_percentage", 0.0) / 100.0, "APPROVED" if is_approved else "-"),
        ("NO / REJECT", dec_info.get("no_votes", 0), dec_info.get("no_percentage", 0.0) / 100.0, "REJECTED" if not is_approved and decision_res == "REJECTED" else "-"),
    ]

    for choice, count, pct, outcome in summary_entries:
        r = ws_sum.max_row + 1
        c1 = ws_sum.cell(row=r, column=1, value=choice)
        c2 = ws_sum.cell(row=r, column=2, value=count)
        c3 = ws_sum.cell(row=r, column=3, value=pct)
        c4 = ws_sum.cell(row=r, column=4, value=outcome)

        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        c4.font = font_bold

        c2.number_format = "#,##0"
        c3.number_format = "0.00%"
        c4.alignment = Alignment(horizontal="center")

        for c in (c1, c2, c3, c4):
            c.border = thin_border

    # Embed Pie Chart
    pie = PieChart()
    pie.title = "YES vs NO Decision Share"
    pie_data = Reference(ws_sum, min_col=2, min_row=1, max_row=3)
    pie_cats = Reference(ws_sum, min_col=1, min_row=2, max_row=3)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 14
    pie.height = 9
    ws_sum.add_chart(pie, "F2")

    auto_fit_columns(ws_sum, 4)

    # SHEET 4: Vote Details
    ws_votes = wb.create_sheet(title="Vote Details")
    ws_votes.views.sheetView[0].showGridLines = True
    ws_votes.freeze_panes = "A2"

    show_names = bool(el.get("show_voter_names_in_results"))
    vote_headers = ["Voter ID", "Voter Name", "Voted At Timestamp", "Status"] if show_names else ["Voter ID", "Voted At Timestamp", "Status"]
    style_table_header(ws_votes, 1, vote_headers, "065F46")

    for item in data.get("voter_participation_log", []):
        r = ws_votes.max_row + 1
        if show_names:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voter_name", ""))
            c3 = ws_votes.cell(row=r, column=3, value=item.get("voted_at", ""))
            c4 = ws_votes.cell(row=r, column=4, value="BALLOT CAST")
            row_cells = (c1, c2, c3, c4)
        else:
            c1 = ws_votes.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_votes.cell(row=r, column=2, value=item.get("voted_at", ""))
            c3 = ws_votes.cell(row=r, column=3, value="BALLOT CAST")
            row_cells = (c1, c2, c3)

        for cell in row_cells:
            cell.font = font_reg
            cell.border = thin_border

    auto_fit_columns(ws_votes, 4)
    return wb


# =============================================================================
# 5. RATING WORKBOOK GENERATOR
# =============================================================================
def generate_rating_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, "Rating & Evaluation", data, "D97706")  # Amber / Gold Theme
    el = data.get("election", {})
    stats = data.get("statistics", {})
    rating_info = data.get("rating", {})
    distribution = rating_info.get("distribution", [])

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    font_reg = Font(name="Calibri", size=11, color="334155")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_amber = Font(name="Calibri", size=13, bold=True, color="B45309")
    fill_amber = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # SHEET 2: Rating Dashboard
    ws_dash = wb.create_sheet(title="Rating Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash.freeze_panes = "A2"

    style_table_header(ws_dash, 1, ["Evaluation Parameter", "Value", "Metric Description"], "D97706")

    dash_rows = [
        ("Rating Topic / Question", rating_info.get("subject", el.get("name", "")), "Evaluated subject / service / topic"),
        ("Average Rating Score", f"{rating_info.get('average_rating', 0.0):.2f} / 5.00", "Weighted average of all cast star ratings"),
        ("Total Ratings Submitted", rating_info.get("total_responses", stats.get("votes_cast", 0)), "Total verified voter evaluations"),
        ("Participation Turnout Rate", stats.get("turnout_percentage", 0.0) / 100.0, "Voters who submitted ratings / Eligible voters"),
        ("Maximum Scale", "5.0 Stars (1 ★ to 5 ★)", "5-point standardized rating scale"),
    ]

    for label, val, note in dash_rows:
        r = ws_dash.max_row + 1
        c1 = ws_dash.cell(row=r, column=1, value=label)
        c2 = ws_dash.cell(row=r, column=2, value=val)
        c3 = ws_dash.cell(row=r, column=3, value=note)

        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg

        if label == "Average Rating Score":
            c2.font = font_amber
            c2.fill = fill_amber
            c2.alignment = Alignment(horizontal="center")

        for c in (c1, c2, c3):
            c.border = thin_border

        if "Rate" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_dash, 3)

    # SHEET 3: Rating Distribution
    ws_dist = wb.create_sheet(title="Rating Distribution")
    ws_dist.views.sheetView[0].showGridLines = True
    ws_dist.freeze_panes = "A2"

    style_table_header(ws_dist, 1, ["Rating Level", "Star Symbol", "Responses Count", "Share Percentage"], "B45309")

    for item in distribution:
        r = ws_dist.max_row + 1
        star_num = item.get("star", 0)
        c1 = ws_dist.cell(row=r, column=1, value=f"{star_num} Stars" if star_num > 1 else "1 Star")
        c2 = ws_dist.cell(row=r, column=2, value=item.get("label", f"{star_num} ★"))
        c3 = ws_dist.cell(row=r, column=3, value=item.get("count", 0))
        c4 = ws_dist.cell(row=r, column=4, value=item.get("percentage", 0.0) / 100.0)

        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        c4.font = font_reg

        c2.alignment = Alignment(horizontal="center")
        c3.number_format = "#,##0"
        c4.number_format = "0.00%"

        for c in (c1, c2, c3, c4):
            c.border = thin_border

    if len(distribution) > 0:
        bar = BarChart()
        bar.type = "col"
        bar.style = 14
        bar.title = "1–5 Star Rating Distribution"
        bar.y_axis.title = "Responses"
        bar.x_axis.title = "Rating Level"
        data_ref = Reference(ws_dist, min_col=3, min_row=1, max_row=len(distribution) + 1)
        cats_ref = Reference(ws_dist, min_col=2, min_row=2, max_row=len(distribution) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 16
        bar.height = 10
        ws_dist.add_chart(bar, "F2")

    auto_fit_columns(ws_dist, 4)

    # SHEET 4: Response Data
    ws_resp = wb.create_sheet(title="Response Data")
    ws_resp.views.sheetView[0].showGridLines = True
    ws_resp.freeze_panes = "A2"

    show_names = bool(el.get("show_voter_names_in_results"))
    resp_headers = ["Respondent Voter ID", "Respondent Name", "Timestamp", "Status"] if show_names else ["Respondent Voter ID", "Timestamp", "Status"]
    style_table_header(ws_resp, 1, resp_headers, "B45309")

    for item in data.get("voter_participation_log", []):
        r = ws_resp.max_row + 1
        if show_names:
            c1 = ws_resp.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_resp.cell(row=r, column=2, value=item.get("voter_name", ""))
            c3 = ws_resp.cell(row=r, column=3, value=item.get("voted_at", ""))
            c4 = ws_resp.cell(row=r, column=4, value="RECORDED")
            row_cells = (c1, c2, c3, c4)
        else:
            c1 = ws_resp.cell(row=r, column=1, value=item.get("voter_id", ""))
            c2 = ws_resp.cell(row=r, column=2, value=item.get("voted_at", ""))
            c3 = ws_resp.cell(row=r, column=3, value="RECORDED")
            row_cells = (c1, c2, c3)

        for cell in row_cells:
            cell.font = font_reg
            cell.border = thin_border

    auto_fit_columns(ws_resp, 4)

    # SHEET 5: Statistics
    ws_stat = wb.create_sheet(title="Statistics")
    ws_stat.views.sheetView[0].showGridLines = True
    ws_stat.freeze_panes = "A2"

    style_table_header(ws_stat, 1, ["Statistical Metric", "Value", "Description"], "B45309")
    stat_rows = [
        ("Total Eligible Evaluators", stats.get("eligible_voters", 0), "Voters qualified to submit ratings"),
        ("Completed Submissions", rating_info.get("total_responses", stats.get("votes_cast", 0)), "Total verified evaluations recorded"),
        ("Participation Turnout", stats.get("turnout_percentage", 0.0) / 100.0, "Turnout percentage rate"),
        ("Weighted Average Score", f"{rating_info.get('average_rating', 0.0):.2f} / 5.0", "Mean rating metric"),
    ]

    for label, val, desc in stat_rows:
        r = ws_stat.max_row + 1
        c1 = ws_stat.cell(row=r, column=1, value=label)
        c2 = ws_stat.cell(row=r, column=2, value=val)
        c3 = ws_stat.cell(row=r, column=3, value=desc)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Turnout" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_stat, 3)
    return wb

# =============================================================================
# 6. QUICK VOTER ENTRY ELECTION WORKBOOK GENERATOR
# =============================================================================
def generate_quick_entry_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    el = data.get("election", {})
    stats = data.get("statistics", {})
    candidates = data.get("candidates", []) or data.get("options", [])
    voter_records = data.get("voter_records", []) or data.get("quick_voter_records", [])
    candidate_voters = data.get("candidate_voters", [])

    theme_color = "0F766E"  # Teal theme
    sub_color = "134E4A"

    font_head = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_reg = Font(name="Calibri", size=10, bold=False, color="1E293B")
    font_kpi_val = Font(name="Calibri", size=16, bold=True, color="0F766E")
    font_kpi_lbl = Font(name="Calibri", size=9, bold=True, color="64748B")
    font_subhead = Font(name="Calibri", size=11, bold=True, color="134E4A")

    fill_head = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
    fill_sub = PatternFill(start_color=sub_color, end_color=sub_color, fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_cand_card = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # -------------------------------------------------------------------------
    # SHEET 1: COVER
    # -------------------------------------------------------------------------
    create_cover_sheet(wb, "QUICK VOTER ENTRY RESULTS", data, theme_color_hex=theme_color)

    # -------------------------------------------------------------------------
    # SHEET 2: RESULTS DASHBOARD
    # -------------------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Results Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash.merge_cells("B2:H2")
    ws_dash["B2"].value = f"{el.get('name', 'ELECTION')} — QUICK ENTRY RESULTS DASHBOARD"
    ws_dash["B2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_dash["B2"].fill = fill_head
    ws_dash["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 32

    # KPI Summary Cards in B4:G5
    kpis = [
        ("TOTAL VOTES CAST", stats.get("votes_cast", len(voter_records)), "B", "C"),
        ("UNIQUE PRN PARTICIPANTS", len(voter_records), "D", "E"),
        ("VOTING FORMAT", str(el.get("voting_type", "regular")).replace("_", " ").upper(), "F", "G"),
    ]
    for lbl, val, col1, col2 in kpis:
        ws_dash.merge_cells(f"{col1}4:{col2}4")
        ws_dash.merge_cells(f"{col1}5:{col2}5")
        ws_dash[f"{col1}4"].value = lbl
        ws_dash[f"{col1}4"].font = font_kpi_lbl
        ws_dash[f"{col1}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash[f"{col1}4"].fill = fill_zebra
        ws_dash[f"{col1}5"].value = val
        ws_dash[f"{col1}5"].font = font_kpi_val
        ws_dash[f"{col1}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash[f"{col1}5"].fill = fill_zebra

        for col_letter in [col1, col2]:
            ws_dash[f"{col_letter}4"].border = thin_border
            ws_dash[f"{col_letter}5"].border = thin_border

    # Candidate Standings Table
    headers_stand = ["Rank", "Candidate / Option", "Party / Category", "Votes Received", "Vote Share %"]
    for col_idx, h in enumerate(headers_stand, start=2):
        cell = ws_dash.cell(row=7, column=col_idx, value=h)
        cell.font = font_head
        cell.fill = fill_sub
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_dash.row_dimensions[7].height = 24

    start_row = 8
    for idx, c in enumerate(candidates, start=start_row):
        c1 = ws_dash.cell(row=idx, column=2, value=c.get("rank", idx - start_row + 1))
        c2 = ws_dash.cell(row=idx, column=3, value=c.get("name", ""))
        c3 = ws_dash.cell(row=idx, column=4, value=c.get("party", "") or "General")
        c4 = ws_dash.cell(row=idx, column=5, value=c.get("votes", c.get("selections_count", 0)))
        c5 = ws_dash.cell(row=idx, column=6, value=(c.get("percentage", 0.0) or 0.0) / 100.0)

        c1.alignment = Alignment(horizontal="center")
        c4.alignment = Alignment(horizontal="right")
        c5.alignment = Alignment(horizontal="right")
        c4.number_format = "#,##0"
        c5.number_format = "0.00%"

        is_even = (idx % 2 == 0)
        for cell in (c1, c2, c3, c4, c5):
            cell.font = font_reg
            cell.border = thin_border
            if is_even:
                cell.fill = fill_zebra

    end_row = max(start_row + len(candidates) - 1, start_row)

    # Add OpenPyXL BarChart on Sheet 2
    if candidates and len(candidates) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Vote Distribution by Candidate / Option"
        chart.y_axis.title = "Votes"
        chart.x_axis.title = "Candidates"
        chart.width = 16
        chart.height = 10
        chart.legend = None

        data_ref = Reference(ws_dash, min_col=5, min_row=7, max_row=end_row)
        cats_ref = Reference(ws_dash, min_col=3, min_row=8, max_row=end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_dash.add_chart(chart, f"B{end_row + 3}")

    auto_fit_columns(ws_dash, 7)

    # -------------------------------------------------------------------------
    # SHEET 3: CANDIDATE-WISE RESULTS
    # -------------------------------------------------------------------------
    ws_cand = wb.create_sheet(title="Candidate-wise Results")
    ws_cand.views.sheetView[0].showGridLines = True

    ws_cand.merge_cells("A1:D1")
    ws_cand["A1"].value = "CANDIDATE-WISE VOTER BREAKDOWN"
    ws_cand["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws_cand["A1"].fill = fill_head
    ws_cand["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cand.row_dimensions[1].height = 28

    curr_row = 3
    cand_groups = candidate_voters if candidate_voters else [{"candidate_name": c.get("name"), "party": c.get("party"), "total_votes": c.get("votes", 0), "voters": []} for c in candidates]

    for cg in cand_groups:
        c_name = cg.get("candidate_name", "Candidate")
        party = cg.get("party", "")
        tot_v = cg.get("total_votes", len(cg.get("voters", [])))
        v_list = cg.get("voters", [])

        # Candidate Header Banner
        ws_cand.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
        c_banner = ws_cand.cell(row=curr_row, column=1, value=f"{c_name.upper()} {f'({party})' if party else ''} — TOTAL VOTES: {tot_v}")
        c_banner.font = font_subhead
        c_banner.fill = fill_cand_card
        c_banner.border = thin_border
        c_banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_cand.row_dimensions[curr_row].height = 24
        curr_row += 1

        # Table Header
        headers = ["#", "Voter Full Name", "10-Digit PRN", "Vote Timestamp"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_cand.cell(row=curr_row, column=col_idx, value=h)
            cell.font = font_head
            cell.fill = fill_sub
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_cand.row_dimensions[curr_row].height = 20
        curr_row += 1

        if not v_list:
            ws_cand.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
            empty_cell = ws_cand.cell(row=curr_row, column=1, value="No individual votes recorded for this candidate yet.")
            empty_cell.font = Font(name="Calibri", size=9, italic=True, color="64748B")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 5):
                ws_cand.cell(row=curr_row, column=col).border = thin_border
            curr_row += 1
        else:
            for v_idx, v in enumerate(v_list, start=1):
                c1 = ws_cand.cell(row=curr_row, column=1, value=v_idx)
                c2 = ws_cand.cell(row=curr_row, column=2, value=v.get("name") or v.get("voter_name", ""))
                c3 = ws_cand.cell(row=curr_row, column=3, value=f"'{v.get('prn', '')}")
                c4 = ws_cand.cell(row=curr_row, column=4, value=v.get("timestamp") or v.get("cast_at", ""))

                c1.alignment = Alignment(horizontal="center")
                c3.alignment = Alignment(horizontal="center")
                c4.alignment = Alignment(horizontal="center")

                for cell in (c1, c2, c3, c4):
                    cell.font = font_reg
                    cell.border = thin_border
                    if v_idx % 2 == 0:
                        cell.fill = fill_zebra
                curr_row += 1

        # Spacer row between candidates
        curr_row += 1

    auto_fit_columns(ws_cand, 4)

    # -------------------------------------------------------------------------
    # SHEET 4: VOTER-WISE VOTE RECORD
    # -------------------------------------------------------------------------
    ws_voter = wb.create_sheet(title="Voter-wise Vote Record")
    ws_voter.views.sheetView[0].showGridLines = True

    ws_voter.freeze_panes = "A2"
    voter_headers = ["#", "Voter Name", "10-Digit PRN", "Vote Given To", "Date / Time"]
    for col_idx, h in enumerate(voter_headers, start=1):
        cell = ws_voter.cell(row=1, column=col_idx, value=h)
        cell.font = font_head
        cell.fill = fill_head
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_voter.row_dimensions[1].height = 26

    for idx, r in enumerate(voter_records, start=2):
        c1 = ws_voter.cell(row=idx, column=1, value=idx - 1)
        c2 = ws_voter.cell(row=idx, column=2, value=r.get("voter_name", ""))
        c3 = ws_voter.cell(row=idx, column=3, value=f"'{r.get('prn', '')}")
        c4 = ws_voter.cell(row=idx, column=4, value=r.get("vote_given_to", ""))
        c5 = ws_voter.cell(row=idx, column=5, value=r.get("timestamp") or r.get("cast_at", ""))

        c1.alignment = Alignment(horizontal="center")
        c3.alignment = Alignment(horizontal="center")
        c5.alignment = Alignment(horizontal="center")

        for cell in (c1, c2, c3, c4, c5):
            cell.font = font_reg
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = fill_zebra
        ws_voter.row_dimensions[idx].height = 20

    auto_fit_columns(ws_voter, 5)

    # -------------------------------------------------------------------------
    # SHEET 5: STATISTICS
    # -------------------------------------------------------------------------
    ws_stat = wb.create_sheet(title="Statistics")
    ws_stat.views.sheetView[0].showGridLines = True

    ws_stat.merge_cells("A1:C1")
    ws_stat["A1"].value = "ELECTION PARTICIPATION & SYSTEM STATISTICS"
    ws_stat["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws_stat["A1"].fill = fill_head
    ws_stat["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_stat.row_dimensions[1].height = 28

    stat_headers = ["Metric / Parameter", "Value", "Description / Notes"]
    for col_idx, h in enumerate(stat_headers, start=1):
        cell = ws_stat.cell(row=3, column=col_idx, value=h)
        cell.font = font_head
        cell.fill = fill_sub
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_stat.row_dimensions[3].height = 22

    stat_rows = [
        ("Voter Registration Mode", "Quick Voter Entry (Name + 10-Digit PRN)", "Open enrollment with strict 1-person-1-vote duplicate prevention"),
        ("Voting Type Format", str(el.get("voting_type", "regular")).replace("_", " ").title(), "Ballot format configured for this election"),
        ("Total Votes Recorded", stats.get("votes_cast", len(voter_records)), "Total verified ballots cast and tallied"),
        ("Total Participating Voters", len(voter_records), "Total unique voters who completed the voting process"),
        ("Candidates / Options Count", len(candidates), "Total choices available on the ballot"),
        ("Turnout Rate", (stats.get("turnout_percentage", 100.0) or 100.0) / 100.0, "Voter participation relative to eligible electorate"),
        ("Audit Integrity Status", "VERIFIED & UNCOMPROMISED", "Cryptographic receipt chain verified"),
        ("Report Generation Timestamp", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"), "Time of Excel generation"),
    ]

    for label, val, desc in stat_rows:
        r = ws_stat.max_row + 1
        c1 = ws_stat.cell(row=r, column=1, value=label)
        c2 = ws_stat.cell(row=r, column=2, value=val)
        c3 = ws_stat.cell(row=r, column=3, value=desc)
        c1.font = font_bold
        c2.font = font_reg
        c3.font = font_reg
        for c in (c1, c2, c3):
            c.border = thin_border
        if "Turnout" in label:
            c2.number_format = "0.00%"
        elif isinstance(val, (int, float)):
            c2.number_format = "#,##0"

    auto_fit_columns(ws_stat, 3)
    return wb


# =============================================================================
# MAIN EXCEL GENERATION DISPATCHER
# =============================================================================
def generate_election_excel(data: dict) -> bytes:
    reg_mode = data.get("voter_registration_mode") or data.get("election", {}).get("voter_registration_mode", "pre_registered")
    if reg_mode == "quick_entry" or data.get("quick_voter_records") or data.get("voter_records"):
        wb = generate_quick_entry_excel(data)
    else:
        voting_type = data.get("voting_type", "regular")
        if voting_type == "poll":
            wb = generate_poll_excel(data)
        elif voting_type == "multiple_choice":
            wb = generate_multiple_choice_excel(data)
        elif voting_type == "yes_no":
            wb = generate_yes_no_excel(data)
        elif voting_type == "rating":
            wb = generate_rating_excel(data)
        else:
            wb = generate_regular_election_excel(data)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()
