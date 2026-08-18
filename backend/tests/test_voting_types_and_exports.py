import io
from datetime import datetime, timedelta, timezone
import openpyxl
from sqlalchemy import select
from app.core.security import create_token, password_hash
from app.models import (
    Candidate,
    Election,
    ElectionState,
    Role,
    User,
    Vote,
    Voter,
    VoterElectionStatus,
)


def _auth_header(user_id: str, role: Role = Role.TEMP_ADMIN) -> dict:
    token = create_token(user_id, role.value, "access", timedelta(minutes=30))
    return {"Authorization": f"Bearer {token}"}


def test_regular_election_results_and_excel(client, db_session):
    now = datetime.now(timezone.utc)
    admin = User(email="regular_admin@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin)
    db_session.flush()

    election = Election(
        name="Student Council Presidential Election 2026",
        election_id="ELEC-REG-001",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        voting_type="regular",
        temp_admin_user_id=admin.id,
    )
    db_session.add(election)
    db_session.flush()

    cand1 = Candidate(election_id=election.id, name="Alice Walker", party="Progressive Alliance", manifesto="Student welfare")
    cand2 = Candidate(election_id=election.id, name="Bob Chen", party="Campus Reform", manifesto="Better campus amenities")
    cand3 = Candidate(election_id=election.id, name="Charlie Davis", party="Independent", manifesto="Open governance")
    db_session.add_all([cand1, cand2, cand3])
    db_session.flush()

    # Cast votes: Alice=5, Bob=3, Charlie=2
    votes = []
    for i in range(5):
        votes.append(Vote(election_id=election.id, candidate_id=cand1.id, receipt_id=f"rec-reg-a-{i}"))
    for i in range(3):
        votes.append(Vote(election_id=election.id, candidate_id=cand2.id, receipt_id=f"rec-reg-b-{i}"))
    for i in range(2):
        votes.append(Vote(election_id=election.id, candidate_id=cand3.id, receipt_id=f"rec-reg-c-{i}"))
    db_session.add_all(votes)
    db_session.commit()

    headers = _auth_header(admin.id)
    # Check results API
    res = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["voting_type"] == "regular"
    assert data["statistics"]["votes_cast"] == 10
    assert data["winner"]["name"] == "Alice Walker"
    assert data["winner"]["votes"] == 5
    assert data["winner"]["percentage"] == 50.0
    assert len(data["candidates"]) == 3
    assert data["candidates"][0]["rank"] == 1

    # Check Excel export
    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200
    assert ".xlsx" in res_excel.headers["content-disposition"]

    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    assert "Cover" in wb.sheetnames
    assert "Results Dashboard" in wb.sheetnames
    assert "Candidate Results" in wb.sheetnames
    assert "Vote Details" in wb.sheetnames
    assert "Voter Statistics" in wb.sheetnames


def test_poll_results_and_excel(client, db_session):
    now = datetime.now(timezone.utc)
    admin = User(email="poll_admin@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin)
    db_session.flush()

    election = Election(
        name="Campus Library Feature Poll",
        election_id="POLL-2026-002",
        description="Which feature should be prioritized for the new campus library?",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        voting_type="poll",
        temp_admin_user_id=admin.id,
    )
    db_session.add(election)
    db_session.flush()

    opt1 = Candidate(election_id=election.id, name="24/7 Study Pods", party="Facilities", manifesto="")
    opt2 = Candidate(election_id=election.id, name="Digital Multimedia Lab", party="Technology", manifesto="")
    opt3 = Candidate(election_id=election.id, name="Outdoor Reading Garden", party="Environment", manifesto="")
    db_session.add_all([opt1, opt2, opt3])
    db_session.flush()

    # Votes: 24/7 Study Pods=8, Digital Lab=4, Reading Garden=3
    votes = []
    for i in range(8):
        votes.append(Vote(election_id=election.id, candidate_id=opt1.id, receipt_id=f"rec-poll-1-{i}"))
    for i in range(4):
        votes.append(Vote(election_id=election.id, candidate_id=opt2.id, receipt_id=f"rec-poll-2-{i}"))
    for i in range(3):
        votes.append(Vote(election_id=election.id, candidate_id=opt3.id, receipt_id=f"rec-poll-3-{i}"))
    db_session.add_all(votes)
    db_session.commit()

    headers = _auth_header(admin.id)
    res = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["voting_type"] == "poll"
    assert "poll" in data
    assert data["poll"]["question"] == "Which feature should be prioritized for the new campus library?"
    assert data["poll"]["total_responses"] == 15
    assert data["poll"]["most_selected_option"]["name"] == "24/7 Study Pods"
    assert data["poll"]["most_selected_option"]["votes"] == 8

    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200
    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    assert "Cover" in wb.sheetnames
    assert "Poll Dashboard" in wb.sheetnames
    assert "Question Results" in wb.sheetnames
    assert "Response Data" in wb.sheetnames
    assert "Statistics" in wb.sheetnames


def test_multiple_choice_results_and_excel(client, db_session):
    now = datetime.now(timezone.utc)
    admin = User(email="mc_admin@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin)
    db_session.flush()

    election = Election(
        name="Sports Committee Board Selection",
        election_id="MC-2026-003",
        description="Select all sports initiatives you would like to support.",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        voting_type="multiple_choice",
        temp_admin_user_id=admin.id,
    )
    db_session.add(election)
    db_session.flush()

    opt1 = Candidate(election_id=election.id, name="Badminton Club", party="", manifesto="")
    opt2 = Candidate(election_id=election.id, name="Basketball League", party="", manifesto="")
    opt3 = Candidate(election_id=election.id, name="Swimming Team", party="", manifesto="")
    db_session.add_all([opt1, opt2, opt3])
    db_session.flush()

    # Simulate 5 voters:
    # Voter 1: opt1, opt2
    # Voter 2: opt1, opt3
    # Voter 3: opt1, opt2, opt3
    # Voter 4: opt2
    # Voter 5: opt1
    # Total voters = 5
    # Total selections = opt1: 4 (80%), opt2: 3 (60%), opt3: 2 (40%) -> Sum = 180% (>100%)
    for v_idx in range(1, 6):
        v_user = User(email=f"voter_mc_{v_idx}@civitas.local", password_hash="h", role=Role.VOTER)
        db_session.add(v_user)
        db_session.flush()
        voter = Voter(
            user_id=v_user.id,
            voter_id=f"VOTER-MC-{v_idx}",
            full_name=f"Voter MC {v_idx}",
            date_of_birth="2000-01-01",
            gender="Other",
            mobile=f"+91980000000{v_idx}",
            address_ciphertext="enc",
            aadhaar_last_four=f"100{v_idx}",
            aadhaar_digest=f"mc_aadhaar_digest_{v_idx}",
        )
        db_session.add(voter)
        db_session.flush()
        st = VoterElectionStatus(voter_id=voter.id, election_id=election.id, eligible=True, voted_at=now)
        db_session.add(st)

    # Add selections
    votes_data = [
        (opt1.id, "r1"), (opt2.id, "r2"), # v1
        (opt1.id, "r3"), (opt3.id, "r4"), # v2
        (opt1.id, "r5"), (opt2.id, "r6"), (opt3.id, "r7"), # v3
        (opt2.id, "r8"), # v4
        (opt1.id, "r9"), # v5
    ]
    for cid, rec in votes_data:
        db_session.add(Vote(election_id=election.id, candidate_id=cid, receipt_id=rec))
    db_session.commit()

    headers = _auth_header(admin.id)
    res = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["voting_type"] == "multiple_choice"
    assert data["multiple_choice"]["total_voters"] == 5
    assert data["multiple_choice"]["total_selections"] == 9
    assert data["multiple_choice"]["average_selections_per_voter"] == 1.8

    # Check that option percentages are based on total participating voters
    opt_map = {o["name"]: o for o in data["candidates"]}
    assert opt_map["Badminton Club"]["selections_count"] == 4
    assert opt_map["Badminton Club"]["percentage_of_voters"] == 80.0
    assert opt_map["Basketball League"]["selections_count"] == 3
    assert opt_map["Basketball League"]["percentage_of_voters"] == 60.0
    assert opt_map["Swimming Team"]["selections_count"] == 2
    assert opt_map["Swimming Team"]["percentage_of_voters"] == 40.0

    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200
    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    assert "Cover" in wb.sheetnames
    assert "Results Dashboard" in wb.sheetnames
    assert "Selection Results" in wb.sheetnames
    assert "Vote Details" in wb.sheetnames
    assert "Statistics" in wb.sheetnames


def test_yes_no_decision_results_and_excel(client, db_session):
    now = datetime.now(timezone.utc)
    admin = User(email="yesno_admin@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin)
    db_session.flush()

    election = Election(
        name="Campus Green Energy Transition Proposal",
        election_id="DEC-2026-004",
        description="Shall CIVITAS allocate 20% of the operational budget to solar power?",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        voting_type="yes_no",
        temp_admin_user_id=admin.id,
    )
    db_session.add(election)
    db_session.flush()

    opt_yes = Candidate(election_id=election.id, name="YES / APPROVE", party="", manifesto="Vote in favor of proposal")
    opt_no = Candidate(election_id=election.id, name="NO / REJECT", party="", manifesto="Vote against proposal")
    db_session.add_all([opt_yes, opt_no])
    db_session.flush()

    # 18 YES votes, 7 NO votes -> 25 total votes -> APPROVED (72.0% vs 28.0%)
    votes = []
    for i in range(18):
        votes.append(Vote(election_id=election.id, candidate_id=opt_yes.id, receipt_id=f"rec-yn-y-{i}"))
    for i in range(7):
        votes.append(Vote(election_id=election.id, candidate_id=opt_no.id, receipt_id=f"rec-yn-n-{i}"))
    db_session.add_all(votes)
    db_session.commit()

    headers = _auth_header(admin.id)
    res = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["voting_type"] == "yes_no"
    assert "decision" in data
    assert data["decision"]["result"] == "APPROVED"
    assert data["decision"]["total_votes"] == 25
    assert data["decision"]["yes_votes"] == 18
    assert data["decision"]["yes_percentage"] == 72.0
    assert data["decision"]["no_votes"] == 7
    assert data["decision"]["no_percentage"] == 28.0
    assert data["decision"]["margin_percentage"] == 44.0

    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200
    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    assert "Cover" in wb.sheetnames
    assert "Decision Dashboard" in wb.sheetnames
    assert "Vote Summary" in wb.sheetnames
    assert "Vote Details" in wb.sheetnames


def test_rating_results_and_excel(client, db_session):
    now = datetime.now(timezone.utc)
    admin = User(email="rating_admin@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin)
    db_session.flush()

    election = Election(
        name="Annual IT Services Satisfaction Survey",
        election_id="RATE-2026-005",
        description="Rate your satisfaction with campus Wi-Fi and IT infrastructure.",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        voting_type="rating",
        temp_admin_user_id=admin.id,
    )
    db_session.add(election)
    db_session.flush()

    stars = []
    for star in [5, 4, 3, 2, 1]:
        c = Candidate(election_id=election.id, name=f"{star} Stars", party=f"{star}", manifesto="")
        stars.append(c)
    db_session.add_all(stars)
    db_session.flush()

    # Ratings cast:
    # 5★: 10
    # 4★: 5
    # 3★: 3
    # 2★: 1
    # 1★: 1
    # Total = 20
    # Weighted Points = 10*5 + 5*4 + 3*3 + 1*2 + 1*1 = 50 + 20 + 9 + 2 + 1 = 82
    # Average = 82 / 20 = 4.10
    cand_by_star = {5: stars[0], 4: stars[1], 3: stars[2], 2: stars[3], 1: stars[4]}
    votes = []
    counts = {5: 10, 4: 5, 3: 3, 2: 1, 1: 1}
    for s, cnt in counts.items():
        for i in range(cnt):
            votes.append(Vote(election_id=election.id, candidate_id=cand_by_star[s].id, receipt_id=f"rec-rate-{s}-{i}"))
    db_session.add_all(votes)
    db_session.commit()

    headers = _auth_header(admin.id)
    res = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["voting_type"] == "rating"
    assert "rating" in data
    assert data["rating"]["average_rating"] == 4.10
    assert data["rating"]["total_responses"] == 20
    assert len(data["rating"]["distribution"]) == 5

    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200
    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    assert "Cover" in wb.sheetnames
    assert "Rating Dashboard" in wb.sheetnames
    assert "Rating Distribution" in wb.sheetnames
    assert "Response Data" in wb.sheetnames
    assert "Statistics" in wb.sheetnames


def test_local_admin_authorization_isolation(client, db_session):
    now = datetime.now(timezone.utc)
    admin_a = User(email="admin_a@civitas.local", password_hash="h", role=Role.TEMP_ADMIN)
    admin_b = User(email="admin_b@civitas.local", password_hash="h", role=Role.TEMP_ADMIN)
    db_session.add_all([admin_a, admin_b])
    db_session.flush()

    election_a = Election(
        name="Admin A Department Election",
        election_id="ELEC-A",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        temp_admin_user_id=admin_a.id,
    )
    election_b = Election(
        name="Admin B Department Election",
        election_id="ELEC-B",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        temp_admin_user_id=admin_b.id,
    )
    db_session.add_all([election_a, election_b])
    db_session.commit()

    headers_a = _auth_header(admin_a.id)
    headers_b = _auth_header(admin_b.id)

    # Admin A accessing Election A -> Allowed 200
    res_a_own = client.get(f"/api/v1/admin/elections/{election_a.id}/results", headers=headers_a)
    assert res_a_own.status_code == 200

    # Admin A attempting to access Election B -> Forbidden 403
    res_a_tamper = client.get(f"/api/v1/admin/elections/{election_b.id}/results", headers=headers_a)
    assert res_a_tamper.status_code == 403

    # Admin A attempting to export Election B Excel -> Forbidden 403
    res_a_export_tamper = client.get(f"/api/v1/admin/elections/{election_b.id}/export/excel", headers=headers_a)
    assert res_a_export_tamper.status_code == 403

    # Admin B accessing Election B -> Allowed 200
    res_b_own = client.get(f"/api/v1/admin/elections/{election_b.id}/results", headers=headers_b)
    assert res_b_own.status_code == 200

    # Admin B attempting to access Election A -> Forbidden 403
    res_b_tamper = client.get(f"/api/v1/admin/elections/{election_a.id}/results", headers=headers_b)
    assert res_b_tamper.status_code == 403

    # GET /admin/elections for Admin A returns ONLY election_a
    list_res_a = client.get("/api/v1/admin/elections", headers=headers_a)
    assert list_res_a.status_code == 200
    elec_ids_a = [e["id"] for e in list_res_a.json()]
    assert str(election_a.id) in elec_ids_a
    assert str(election_b.id) not in elec_ids_a
