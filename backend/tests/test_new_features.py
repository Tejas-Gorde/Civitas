import io
import os
from datetime import datetime, timezone, timedelta
import openpyxl
from app.models import AuthSession, AuthStage, Candidate, Election, ElectionState, Role, User, Vote, Voter
from app.core.security import create_token


def test_voter_photo_storage(client, db_session):
    now = datetime.now(timezone.utc)
    user = User(email="photovoter@civitas.local", password_hash="test", role=Role.VOTER)
    db_session.add(user)
    db_session.flush()

    voter = Voter(
        user_id=user.id,
        voter_id="VOTER-TEST-99",
        full_name="Photo Test Voter",
        date_of_birth="1990-01-01",
        gender="Other",
        mobile="+1555998877",
        address_ciphertext="test",
        aadhaar_last_four="9999",
        aadhaar_digest="digest9999",
    )
    db_session.add(voter)

    election = Election(
        name="Photo Test Election",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
    )
    db_session.add(election)
    db_session.flush()

    auth_session = AuthSession(
        voter_id=voter.id,
        election_id=election.id,
        stage=AuthStage.FINGERPRINT,
        expires_at=now + timedelta(hours=1),
    )
    db_session.add(auth_session)
    db_session.commit()

    # Perform POST /api/v1/verification/photo upload
    img_bytes = b"\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x01\x00`\x00`\x00\x00\xff\xdb\x00C\x00"  # JPEG header
    response = client.post(
        "/api/v1/verification/photo",
        data={"session_id": str(auth_session.id), "voter_id": voter.voter_id, "election_id": election.id},
        files={"file": ("my_camera.jpg", img_bytes, "image/jpeg")},
    )

    assert response.status_code == 200, response.text
    res_data = response.json()
    assert res_data["success"] is True
    assert res_data["message"] == "Voter photo saved successfully"
    assert "photo_id" in res_data
    assert "filename" in res_data

    filename = res_data["filename"]
    assert filename.startswith("VOTER-TEST-99_")
    assert filename.endswith(".jpg")

    # Verify file saved on disk in ~/Desktop/Civitas_Voter_Photos/
    from pathlib import Path

    desktop_dir = Path.home() / "Desktop" / "Civitas_Voter_Photos"
    target_path = desktop_dir / filename
    assert target_path.exists()

    # Clean up test photo file
    if target_path.exists():
        target_path.unlink()



def test_election_results_api_and_excel_export(client, db_session):
    now = datetime.now(timezone.utc)
    
    # Create admin user and token
    admin_user = User(email="admin_test@civitas.local", password_hash="hash", role=Role.ADMIN)
    db_session.add(admin_user)
    db_session.flush()
    token = create_token(admin_user.id, Role.ADMIN.value, "access", timedelta(minutes=15))
    headers = {"Authorization": f"Bearer {token}"}


    election = Election(
        name="Excel Test Election 2026",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
    )
    db_session.add(election)
    db_session.flush()

    cand1 = Candidate(election_id=election.id, name="Candidate Alpha", party="Alpha Party", manifesto="Plan A")
    cand2 = Candidate(election_id=election.id, name="Candidate Beta", party="Beta Party", manifesto="Plan B")
    db_session.add_all([cand1, cand2])
    db_session.flush()

    vote1 = Vote(election_id=election.id, candidate_id=cand1.id, receipt_id="receipt-001")
    vote2 = Vote(election_id=election.id, candidate_id=cand1.id, receipt_id="receipt-002")
    vote3 = Vote(election_id=election.id, candidate_id=cand2.id, receipt_id="receipt-003")
    db_session.add_all([vote1, vote2, vote3])
    db_session.commit()

    # 1. Test GET /api/v1/admin/elections/{id}/results
    res_api = client.get(f"/api/v1/admin/elections/{election.id}/results", headers=headers)
    assert res_api.status_code == 200, res_api.text
    results_json = res_api.json()

    assert results_json["election"]["name"] == "Excel Test Election 2026"
    assert results_json["statistics"]["votes_cast"] == 3
    assert len(results_json["candidates"]) == 2

    # Winner candidate Alpha should be rank 1 with 2 votes
    top_cand = results_json["candidates"][0]
    assert top_cand["name"] == "Candidate Alpha"
    assert top_cand["votes"] == 2
    assert top_cand["rank"] == 1

    # 2. Test GET /api/v1/admin/elections/{id}/export/excel
    res_excel = client.get(f"/api/v1/admin/elections/{election.id}/export/excel", headers=headers)
    assert res_excel.status_code == 200, res_excel.text
    assert "Civitas_Excel_Test_Election_2026_Results" in res_excel.headers["content-disposition"]
    assert ".xlsx" in res_excel.headers["content-disposition"]


    # Parse Excel workbook binary and check worksheets
    wb = openpyxl.load_workbook(filename=io.BytesIO(res_excel.content))
    sheet_names = wb.sheetnames
    assert "Cover" in sheet_names
    assert "Results Dashboard" in sheet_names
    assert "Candidate Results" in sheet_names
    assert "Vote Details" in sheet_names
    assert "Voter Statistics" in sheet_names

    ws_cand = wb["Candidate Results"]
    assert ws_cand.cell(row=1, column=1).value == "Rank"
    assert ws_cand.cell(row=2, column=2).value == "Candidate Alpha"
    assert ws_cand.cell(row=2, column=4).value == 2


def test_election_specific_voter_management(client, db_session):
    now = datetime.now(timezone.utc)

    # Admin user setup as Local Admin
    admin_user = User(email="admin_voters@civitas.local", password_hash="hash", role=Role.TEMP_ADMIN)
    db_session.add(admin_user)
    db_session.flush()
    token = create_token(admin_user.id, Role.TEMP_ADMIN.value, "access", timedelta(minutes=15))
    headers = {"Authorization": f"Bearer {token}"}

    # Create Election A and Election B assigned to this Local Admin
    election_a = Election(
        name="General Election A",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        temp_admin_user_id=admin_user.id,
    )
    election_b = Election(
        name="Student Election B",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        temp_admin_user_id=admin_user.id,
    )
    election_c = Election(
        name="Civic Election C",
        starts_at=now - timedelta(days=1),
        ends_at=now + timedelta(days=1),
        state=ElectionState.OPEN,
        temp_admin_user_id=admin_user.id,
    )
    db_session.add_all([election_a, election_b, election_c])
    db_session.commit()

    # TEST 1: Register Tejas Gorde (VOTER-1001) for Election A
    res_add_a = client.post(
        f"/api/v1/admin/elections/{election_a.id}/voters",
        headers=headers,
        json={
            "full_name": "Tejas Gorde",
            "voter_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
            "email": "tejas@example.com",
            "mobile": "+15551001001",
            "is_eligible": True,
        },
    )
    assert res_add_a.status_code == 201, res_add_a.text
    assert res_add_a.json()["success"] is True

    # Verify eligibility for Election A
    res_verify_a = client.post(
        "/api/v1/voting/verify-voter",
        json={
            "election_id": str(election_a.id),
            "voter_registration_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
        },
    )
    assert res_verify_a.status_code == 200, res_verify_a.text
    assert res_verify_a.json()["eligible"] is True

    # TEST 2: Verify voter is NOT eligible for Election B
    res_verify_b_fail = client.post(
        "/api/v1/voting/verify-voter",
        json={
            "election_id": str(election_b.id),
            "voter_registration_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
        },
    )
    assert res_verify_b_fail.status_code == 403

    # TEST 3: Add the SAME voter (VOTER-1001) to Election B
    voters_before_count = db_session.query(Voter).count()
    res_add_b = client.post(
        f"/api/v1/admin/elections/{election_b.id}/voters",
        headers=headers,
        json={
            "full_name": "Tejas Gorde",
            "voter_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
            "email": "tejas@example.com",
            "mobile": "+15551001001",
            "is_eligible": True,
        },
    )
    assert res_add_b.status_code == 201, res_add_b.text
    assert res_add_b.json()["success"] is True
    # Ensure no duplicate Voter identity row was created in `voters` table
    voters_after_count = db_session.query(Voter).count()
    assert voters_after_count == voters_before_count

    # TEST 4: Try adding the same voter again to Election B -> should fail with 409
    res_add_b_dup = client.post(
        f"/api/v1/admin/elections/{election_b.id}/voters",
        headers=headers,
        json={
            "full_name": "Tejas Gorde",
            "voter_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
            "email": "tejas@example.com",
            "mobile": "+15551001001",
            "is_eligible": True,
        },
    )
    assert res_add_b_dup.status_code == 409
    assert "already registered" in res_add_b_dup.json()["detail"].lower()

    # TEST 5 & 6: Verify voter is eligible for both Election A and Election B
    res_verify_a_again = client.post(
        "/api/v1/voting/verify-voter",
        json={
            "election_id": str(election_a.id),
            "voter_registration_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
        },
    )
    assert res_verify_a_again.status_code == 200

    res_verify_b_now = client.post(
        "/api/v1/voting/verify-voter",
        json={
            "election_id": str(election_b.id),
            "voter_registration_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
        },
    )
    assert res_verify_b_now.status_code == 200

    # TEST 7: Verify voter is NOT eligible for Election C
    res_verify_c_fail = client.post(
        "/api/v1/voting/verify-voter",
        json={
            "election_id": str(election_c.id),
            "voter_registration_id": "VOTER-1001",
            "voter_password": "VoterPassword123!",
        },
    )
    assert res_verify_c_fail.status_code == 403

